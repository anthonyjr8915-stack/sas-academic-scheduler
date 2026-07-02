"""
Excel import/export — the *only* place Excel touches the system.

Per the architecture decision in the SRS: Excel is an interchange format, never
the source of truth. Import turns a workbook into an in-memory `Problem`; export
turns a `Solution` into a workbook of printable timetables. The database (added
later) sits in between; nothing here reads or writes it.

Import workbook layout (see `write_template` for a generated example):
  Config  : key | value        (days, periods_per_day, lunch_period, assembly_mon)
  Teachers: id | name | max_per_day | unavailable_days
  Classes : id | name | class_teacher_id
  Labs    : kind | capacity
  Plan    : klass_id | subject | teacher_id | per_week | double_blocks | lab_kind | preference
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.scheduler.engine import Solution
from app.scheduler.problem import (
    Klass,
    LabPool,
    Lesson,
    Preference,
    Problem,
    SolverConfig,
    Teacher,
    TimeGrid,
)


# --------------------------------------------------------------------- import
def _rows(ws) -> list[dict]:
    """Read a sheet with a header row into a list of dict rows (blank rows skipped)."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for raw in rows[1:]:
        if all(c is None or str(c).strip() == "" for c in raw):
            continue
        out.append({headers[i]: raw[i] for i in range(len(headers)) if headers[i]})
    return out


def read_problem(path: str | Path) -> Problem:
    wb = load_workbook(path, data_only=True)

    cfg = {r["key"]: r["value"] for r in _rows(wb["Config"])}
    days = [d.strip() for d in str(cfg["days"]).split(",")]
    periods = int(cfg["periods_per_day"])
    grid = TimeGrid(days=days, periods_per_day=periods)

    global_blocked: set[int] = set()
    if cfg.get("lunch_period"):
        lp = int(cfg["lunch_period"]) - 1  # 1-based in sheet
        global_blocked |= {grid.slot(d, lp) for d in range(grid.num_days)}
    if str(cfg.get("assembly_mon", "")).strip().lower() in ("1", "true", "yes"):
        global_blocked |= {grid.slot(0, 0)}

    day_index = {name.lower(): i for i, name in enumerate(days)}

    teachers = []
    for r in _rows(wb["Teachers"]):
        unavailable: set[int] = set()
        for token in str(r.get("unavailable_days") or "").split(","):
            token = token.strip().lower()
            if token in day_index:
                d = day_index[token]
                unavailable |= set(grid.slots_on_day(d))
        teachers.append(Teacher(
            id=str(r["id"]).strip(),
            name=str(r["name"]).strip(),
            max_periods_per_day=int(r["max_per_day"]) if r.get("max_per_day") else None,
            unavailable_slots=unavailable,
        ))

    classes = [
        Klass(
            id=str(r["id"]).strip(),
            name=str(r["name"]).strip(),
            class_teacher_id=str(r["class_teacher_id"]).strip() if r.get("class_teacher_id") else None,
        )
        for r in _rows(wb["Classes"])
    ]

    lab_pools = [
        LabPool(kind=str(r["kind"]).strip(), capacity=int(r["capacity"]))
        for r in _rows(wb["Labs"])
    ] if "Labs" in wb.sheetnames else []

    lessons: list[Lesson] = []
    for r in _rows(wb["Plan"]):
        klass_id = str(r["klass_id"]).strip()
        subject = str(r["subject"]).strip()
        teacher_id = str(r["teacher_id"]).strip()
        per_week = int(r["per_week"])
        doubles = int(r.get("double_blocks") or 0)
        lab_kind = str(r["lab_kind"]).strip() if r.get("lab_kind") else None
        pref = Preference(str(r.get("preference") or "none").strip().lower())

        n, idx = per_week, 0
        for _ in range(doubles):
            if n >= 2:
                lessons.append(Lesson(
                    id=f"{klass_id}_{subject}_{idx}", klass_id=klass_id, subject=subject,
                    teacher_id=teacher_id, length=2, lab_kind=lab_kind, preference=pref))
                n -= 2
                idx += 1
        for _ in range(n):
            lessons.append(Lesson(
                id=f"{klass_id}_{subject}_{idx}", klass_id=klass_id, subject=subject,
                teacher_id=teacher_id, length=1, lab_kind=lab_kind, preference=pref))
            idx += 1

    return Problem(
        grid=grid, teachers=teachers, classes=classes, lessons=lessons,
        lab_pools=lab_pools, global_blocked_slots=global_blocked,
        config=SolverConfig(),
    )


# --------------------------------------------------------------------- export
_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_BREAK_FILL = PatternFill("solid", fgColor="E5E7EB")
_LAB_FILL = PatternFill("solid", fgColor="FEF3C7")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize(ws, ncols: int, width: int = 14) -> None:
    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = width


def write_solution(problem: Problem, solution: Solution, path: str | Path) -> Path:
    """Write per-class and per-teacher timetables plus a score sheet."""
    grid = problem.grid
    wb = Workbook()
    wb.remove(wb.active)

    def grid_sheet(title: str, cell_of, blocked_of):
        ws = wb.create_sheet(title[:31])
        ws.cell(1, 1, "Period").font = _HEADER_FONT
        ws.cell(1, 1).fill = _HEADER_FILL
        ws.cell(1, 1).alignment = _CENTER
        for d, day in enumerate(grid.days):
            c = ws.cell(1, 2 + d, day)
            c.font, c.fill, c.alignment = _HEADER_FONT, _HEADER_FILL, _CENTER
        for period in range(grid.periods_per_day):
            ws.cell(2 + period, 1, f"P{period + 1}").font = Font(bold=True)
            for d in range(grid.num_days):
                slot = grid.slot(d, period)
                cell = ws.cell(2 + period, 2 + d)
                cell.alignment = _CENTER
                if blocked_of(slot):
                    cell.value = "break"
                    cell.fill = _BREAK_FILL
                else:
                    text, is_lab = cell_of(slot)
                    cell.value = text
                    if is_lab:
                        cell.fill = _LAB_FILL
        _autosize(ws, grid.num_days + 1)
        return ws

    # Per-class sheets.
    by_class: dict[str, dict[int, tuple[str, bool]]] = {c.id: {} for c in problem.classes}
    by_teacher: dict[str, dict[int, tuple[str, bool]]] = {t.id: {} for t in problem.teachers}
    for pl in solution.placements:
        for k in range(pl.length):
            slot = pl.start_slot + k
            label = pl.subject + ("+" if k else "")
            by_class[pl.klass_id][slot] = (label, bool(pl.lab_kind))
            by_teacher[pl.teacher_id][slot] = (f"{pl.klass_id}:{pl.subject}", bool(pl.lab_kind))

    for klass in problem.classes:
        blocked = problem.global_blocked_slots | klass.blocked_slots
        grid_sheet(
            f"Class {klass.name}",
            lambda s, m=by_class[klass.id]: m.get(s, ("", False)),
            lambda s, b=blocked: s in b,
        )
    for teacher in problem.teachers:
        grid_sheet(
            f"T {teacher.name}",
            lambda s, m=by_teacher[teacher.id]: m.get(s, ("", False)),
            lambda s, t=teacher: s in (problem.global_blocked_slots | t.unavailable_slots),
        )

    # Score sheet.
    ws = wb.create_sheet("Score")
    ws.append(["Optimization score", solution.score])
    ws.append(["Solver status", solution.status])
    ws.append([])
    ws.append(["Rule", "Points", "Detail"])
    for c in range(1, 4):
        ws.cell(4, c).font = _HEADER_FONT
        ws.cell(4, c).fill = _HEADER_FILL
    for line in solution.score_breakdown:
        ws.append([line.rule, line.points, line.detail])
    _autosize(ws, 3, 26)

    path = Path(path)
    wb.save(path)
    return path


# ------------------------------------------------------------------- template
def write_template(path: str | Path) -> Path:
    """Emit a ready-to-fill import workbook so users know the exact format."""
    wb = Workbook()
    wb.remove(wb.active)

    def sheet(name, headers, rows):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            ws.cell(1, c).font = _HEADER_FONT
            ws.cell(1, c).fill = _HEADER_FILL
        for r in rows:
            ws.append(r)
        _autosize(ws, len(headers))

    sheet("Config", ["key", "value"], [
        ["days", "Mon,Tue,Wed,Thu,Fri"],
        ["periods_per_day", 7],
        ["lunch_period", 4],
        ["assembly_mon", "yes"],
    ])
    sheet("Teachers", ["id", "name", "max_per_day", "unavailable_days"], [
        ["t_maths", "Rao", 6, ""],
        ["t_sci", "Iyer", 6, ""],
        ["t_eng", "Khan", 6, ""],
        ["t_hin", "Verma", 6, "Thu,Fri"],
    ])
    sheet("Classes", ["id", "name", "class_teacher_id"], [
        ["IX_A", "IX-A", "t_maths"],
    ])
    sheet("Labs", ["kind", "capacity"], [
        ["science_lab", 1],
        ["computer_lab", 1],
    ])
    sheet("Plan", ["klass_id", "subject", "teacher_id", "per_week",
                   "double_blocks", "lab_kind", "preference"], [
        ["IX_A", "Maths", "t_maths", 6, 0, "", "morning"],
        ["IX_A", "Science", "t_sci", 5, 0, "", "morning"],
        ["IX_A", "SciLab", "t_sci", 2, 1, "science_lab", "none"],
        ["IX_A", "English", "t_eng", 6, 0, "", "none"],
        ["IX_A", "Hindi", "t_hin", 4, 0, "", "none"],
    ])

    path = Path(path)
    wb.save(path)
    return path
