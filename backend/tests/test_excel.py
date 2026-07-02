"""Excel is the interchange layer: a generated template must import back into a
solvable problem, and a solution must export to a readable workbook."""
from __future__ import annotations

from openpyxl import load_workbook

from app.io import excel
from app.scheduler.engine import TimetableEngine
from app.scheduler.render import verify
from demo.demo_data import build_problem


def test_template_roundtrips_into_a_solvable_problem(tmp_path):
    template = tmp_path / "template.xlsx"
    excel.write_template(template)
    problem = excel.read_problem(template)
    assert problem.classes and problem.teachers and problem.lessons
    problem.config.max_seconds = 10.0
    solution = TimetableEngine(problem).solve()
    assert solution.ok, solution.diagnosis
    assert verify(problem, solution) == []


def test_solution_exports_to_workbook(tmp_path):
    problem = build_problem()
    problem.config.max_seconds = 10.0
    solution = TimetableEngine(problem).solve()
    out = tmp_path / "timetable.xlsx"
    excel.write_solution(problem, solution, out)
    wb = load_workbook(out)
    # A sheet per class, per teacher, plus the Score sheet.
    assert any(name.startswith("Class ") for name in wb.sheetnames)
    assert "Score" in wb.sheetnames
